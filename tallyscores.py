from openpyxl.styles import PatternFill, Font

def color_fill(ws, score, row_num):
  score_split = score.split(' ')

  # cycle through four pick columns; odd = common, even = spread
  for col in range(5, 9):
    pick_cell = ws.cell(row=row_num, column=col)
    pick_cell_text = pick_cell.value

    # common pick
    if col % 2 == 1:
      pick_cell_split = pick_cell_text.split(' ')

      # if pick team == winning team, color cell green
      if pick_cell_split[0].upper() == score_split[0]:
        pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        spread = score_split[1].split('-')

        # if game spread == pick spread, spread font == yellow
        if int(spread[0]) - int(spread[1]) == int(pick_cell_split[2]):
          pick_cell.font = Font(color="FFFB00") # yellow
      # else color cell red    
      else:
        pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red

    # TODO - against line
    else:
      line = ws.cell(row=row_num, column=2).value.upper().split(' -')
      spread = score_split[1].split('-')
      print(score_split)
      print(line)
      print(spread)

      try:
        line_num = int(line[1])
      except ValueError:
        line_num = float(line[1][:-1] + '.5')

      if score_split[0] == line[0] and (int(spread[0]) - int(spread[1])) > line_num:
        team = score_split[0]
        if team == pick_cell_text:
          pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        else:
          pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
      elif score_split[0] != line[0]:
        team = score_split[0]
        if team == pick_cell_text:
          pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        else: 
          pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red
      else:
        if pick_cell_text != line[0]:
          pick_cell.fill = PatternFill("solid", fgColor="009051") # green
        else:
          pick_cell.fill = PatternFill("solid", fgColor="FF7E79") # red